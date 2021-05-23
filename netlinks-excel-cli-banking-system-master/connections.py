import os 
import sys 
import openpyxl
import pyinputplus as pyip
import messages

try:
    bank_data = openpyxl.load_workbook("bank_data.xlsx") 
    sheet = bank_data.active  
    rows = sheet.max_row
    cols = sheet.max_column
    columns = {}
    for col_num in range(1, cols+1):
        columns[f"{sheet.cell(row=1, column=col_num).value}"] = col_num

except (FileNotFoundError) as error:
    print(f"File Not Found! {error}")
    bank_data = openpyxl.Workbook()
    sheet = bank_data.active  
    sheet['A1'] = "ID"
    sheet['B1'] = "Name"
    sheet['C1'] = "Address"
    sheet['D1'] = "Balance"
    sheet['E1'] = "Email"
    sheet['F1'] = "Password"
    bank_data.save("bank_data.xlsx")
    print("We Have Created New File Please Restart The Program")
    sys.exit()
